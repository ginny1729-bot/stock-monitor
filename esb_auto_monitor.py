#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
興櫃自動追蹤 - 全自動版
功能：
1. 抓取櫃買中心興櫃股票當日行情表
2. 累積歷史資料 CSV
3. 計算強勢股 / 爆量股 / 自選清單訊號
4. 更新 Excel 追蹤表
5. 推送 LINE 訊息

使用前請先：
- pip install requests pandas openpyxl python-dotenv lxml html5lib
- 複製 .env.example 為 .env，填入 LINE token / user id
"""

from __future__ import annotations
import os
import re
import json
from pathlib import Path
from datetime import datetime
from typing import List, Dict

import pandas as pd
import requests
from openpyxl import load_workbook
from dotenv import load_dotenv

QUOTE_URL = "https://www.tpex.org.tw/zh-tw/esb/trading/info/pricing.html"
STATS_URL = "https://www.tpex.org.tw/zh-tw/esb/psb/trading/statistics/day.html"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
}

def load_config() -> dict:
    load_dotenv()
    watchlist_codes = os.getenv("WATCHLIST_CODES", "7822,6879,4172")
    return {
        "line_token": os.getenv("LINE_CHANNEL_ACCESS_TOKEN", "").strip(),
        "line_user_id": os.getenv("LINE_USER_ID", "").strip(),
        "watchlist_codes": [x.strip() for x in watchlist_codes.split(",") if x.strip()],
        "price_jump_pct": float(os.getenv("PRICE_JUMP_PCT", "5")),
        "volume_surge_multiple": float(os.getenv("VOLUME_SURGE_MULTIPLE", "2")),
        "report_folder": os.getenv("REPORT_FOLDER", "./output"),
        "history_csv": os.getenv("HISTORY_CSV", "./output/esb_history.csv"),
        "workbook_path": os.getenv("WORKBOOK_PATH", "./興櫃自動追蹤_全自動版.xlsx"),
    }

def clean_num(v):
    if pd.isna(v):
        return None
    s = str(v).strip()
    if s in {"", "--", "---", "除權", "除息", "暫停交易"}:
        return None
    s = s.replace(",", "").replace("%", "")
    s = s.replace("△", "").replace("▲", "").replace("▽", "-").replace("▼", "-")
    s = re.sub(r"[^\d\.\-]", "", s)
    if s in {"", "-", ".", "-."}:
        return None
    try:
        return float(s)
    except Exception:
        return None

def try_read_tables(url: str) -> List[pd.DataFrame]:
    html = requests.get(url, headers=HEADERS, timeout=30).text
    return pd.read_html(html)

def detect_quote_table(tables: List[pd.DataFrame]) -> pd.DataFrame:
    for df in tables:
        cols = [str(c) for c in df.columns]
        joined = "|".join(cols)
        if ("代號" in joined or "股票代號" in joined) and ("名稱" in joined or "公司名稱" in joined):
            return df.copy()
    raise RuntimeError("找不到興櫃報價表，可能是櫃買中心欄位名稱變更。")

def normalize_quote_df(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).replace("\n", "").strip() for c in df.columns]
    rename_map = {}
    for c in df.columns:
        if "代號" in c:
            rename_map[c] = "代號"
        elif "名稱" in c:
            rename_map[c] = "名稱"
        elif "開盤" in c:
            rename_map[c] = "開盤"
        elif "最高" in c:
            rename_map[c] = "最高"
        elif "最低" in c:
            rename_map[c] = "最低"
        elif "漲幅" in c:
            rename_map[c] = "漲幅%"
        elif "成交數量" in c or "成交股數" in c or c == "成交":
            rename_map[c] = "成交量"
        elif "最新成交價" in c or "成交均價" in c or "日均價" in c or "收盤" in c:
            rename_map[c] = "最新價"
    df = df.rename(columns=rename_map)

    required = ["代號", "名稱"]
    for col in required:
        if col not in df.columns:
            raise RuntimeError(f"缺少必要欄位：{col}")

    for col in ["開盤","最高","最低","最新價","漲幅%","成交量"]:
        if col not in df.columns:
            df[col] = None
        df[col] = df[col].apply(clean_num)

    df["代號"] = df["代號"].astype(str).str.extract(r"(\d+)")[0]
    df["名稱"] = df["名稱"].astype(str).str.strip()
    df = df.dropna(subset=["代號"])
    df["資料日期"] = datetime.now().strftime("%Y-%m-%d")
    df["來源"] = QUOTE_URL
    return df[["代號","名稱","開盤","最高","最低","最新價","漲幅%","成交量","資料日期","來源"]].drop_duplicates(subset=["代號"])

def update_history(df: pd.DataFrame, history_csv: str) -> pd.DataFrame:
    history_path = Path(history_csv)
    history_path.parent.mkdir(parents=True, exist_ok=True)
    if history_path.exists():
        old = pd.read_csv(history_path, dtype={"代號": str})
        history = pd.concat([old, df], ignore_index=True)
    else:
        history = df.copy()
    history = history.drop_duplicates(subset=["代號","資料日期"], keep="last")
    history = history.sort_values(["代號","資料日期"])
    history.to_csv(history_path, index=False, encoding="utf-8-sig")
    return history

def build_signals(today_df: pd.DataFrame, history_df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    hist = history_df.copy()
    hist["成交量"] = pd.to_numeric(hist["成交量"], errors="coerce")
    hist["資料日期"] = pd.to_datetime(hist["資料日期"], errors="coerce")
    hist = hist.sort_values(["代號","資料日期"])
    hist["5次平均量"] = hist.groupby("代號")["成交量"].transform(lambda s: s.shift(1).rolling(5, min_periods=1).mean())

    latest_hist = hist[hist["資料日期"] == hist["資料日期"].max()][["代號","5次平均量"]]
    df = today_df.merge(latest_hist, on="代號", how="left")
    df["5次平均量"] = pd.to_numeric(df["5次平均量"], errors="coerce")
    df["量比"] = df["成交量"] / df["5次平均量"]

    def decide(row):
        reasons = []
        tags = []
        if pd.notna(row.get("漲幅%")) and row["漲幅%"] >= cfg["price_jump_pct"]:
            tags.append("強勢股")
            reasons.append(f"漲幅 {row['漲幅%']:.1f}%")
        if pd.notna(row.get("量比")) and row["量比"] >= cfg["volume_surge_multiple"]:
            tags.append("爆量股")
            reasons.append(f"量比 {row['量比']:.2f}x")
        if row["代號"] in cfg["watchlist_codes"]:
            tags.append("自選股")
        if not tags:
            return "", ""
        return " / ".join(dict.fromkeys(tags)), "，".join(reasons) if reasons else "自選追蹤"

    signal_cols = df.apply(decide, axis=1, result_type="expand")
    df["訊號"] = signal_cols[0]
    df["原因"] = signal_cols[1]
    signals = df[df["訊號"] != ""].copy()
    signals = signals[["代號","名稱","最新價","漲幅%","成交量","5次平均量","訊號","原因"]]
    signals = signals.sort_values(["訊號","漲幅%"], ascending=[True, False])
    return signals

def send_line(token: str, user_id: str, text: str) -> bool:
    if not token or not user_id:
        return False
    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    payload = {
        "to": user_id,
        "messages": [{"type": "text", "text": text[:4900]}],
    }
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    r.raise_for_status()
    return True

def build_line_message(signals: pd.DataFrame) -> str:
    if signals.empty:
        return "今日興櫃監控：目前沒有符合條件的強勢/爆量訊號。"
    lines = ["🔥 興櫃盤中監控"]
    for _, row in signals.head(12).iterrows():
        price = "-" if pd.isna(row["最新價"]) else f'{row["最新價"]:.2f}'
        pct = "-" if pd.isna(row["漲幅%"]) else f'{row["漲幅%"]:.1f}%'
        lines.append(f'{row["代號"]} {row["名稱"]} | {row["訊號"]} | 價格 {price} | 漲幅 {pct} | {row["原因"]}')
    return "\n".join(lines)

def write_excel(workbook_path: str, today_df: pd.DataFrame, signals: pd.DataFrame, pushed: bool, note: str):
    wb = load_workbook(workbook_path)
    ws_raw = wb["原始報價"]
    ws_sig = wb["訊號清單"]
    ws_log = wb["執行紀錄"]

    def clear_sheet_data(ws, start_row=2):
        if ws.max_row >= start_row:
            ws.delete_rows(start_row, ws.max_row - start_row + 1)

    clear_sheet_data(ws_raw, 2)
    clear_sheet_data(ws_sig, 2)

    for row in today_df.itertuples(index=False):
        ws_raw.append(list(row))
    for row in signals.itertuples(index=False):
        ws_sig.append(list(row))

    ws_log.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        int(len(today_df)),
        int(len(signals)),
        "成功" if pushed else "未推播",
        note,
    ])
    wb.save(workbook_path)

def main():
    cfg = load_config()
    Path(cfg["report_folder"]).mkdir(parents=True, exist_ok=True)

    tables = try_read_tables(QUOTE_URL)
    today_df = normalize_quote_df(detect_quote_table(tables))
    history_df = update_history(today_df, cfg["history_csv"])
    signals = build_signals(today_df, history_df, cfg)

    message = build_line_message(signals)
    pushed = False
    note = "完成"
    try:
        pushed = send_line(cfg["line_token"], cfg["line_user_id"], message)
    except Exception as e:
        note = f"LINE 推播失敗：{e}"

    write_excel(cfg["workbook_path"], today_df, signals, pushed, note)

    out_json = Path(cfg["report_folder"]) / f"signals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    out_json.write_text(signals.to_json(orient="records", force_ascii=False, indent=2), encoding="utf-8")

    print("完成：")
    print(f"- 今日報價筆數：{len(today_df)}")
    print(f"- 訊號筆數：{len(signals)}")
    print(f"- Excel：{cfg['workbook_path']}")
    print(f"- History CSV：{cfg['history_csv']}")
    print(f"- JSON：{out_json}")

if __name__ == "__main__":
    main()
